import logging
import requests
from telegram import Update, Bot
from telegram.ext import Application, CommandHandler, MessageHandler, filters, ContextTypes
import google.generativeai as genai
from openpyxl import Workbook, load_workbook
import os
import base64

# Bot and Gemini AI settings
TELEGRAM_TOKEN = '7927390418:AAHxie5HOrtfw3Dlfdj3weegsbQoCydudT8'
GEMINI_API_KEY = 'AIzaSyDI38Ta-tPVgRMiXcafqbTIrk1xwLzNi3k'
EXCEL_FILE = 'resultados.xlsx'

# Gemini API configuration
genai.configure(api_key=GEMINI_API_KEY)
model = genai.GenerativeModel("gemini-1.5-flash")

# Logging configuration
logging.basicConfig(format='%(asctime)s - %(name)s - %(levelname)s - %(message)s', level=logging.INFO)
logger = logging.getLogger(__name__)

# Function to start the bot
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text('Send a photo to analyze!')

# Function to process the received photo
async def handle_photo(update: Update, context: ContextTypes.DEFAULT_TYPE):
    file = await update.message.photo[-1].get_file()
    photo_data = requests.get(file.file_path).content

    # Convert the image to base64
    image_data = base64.b64encode(photo_data).decode('utf-8')

    # Send the photo to Gemini AI
    prompt = "Analyze this photo and give me the total spent!"
    inputs = [{'mime_type': 'image/jpeg', 'data': image_data}, prompt]
    try:
        response = model.generate_content(inputs)
        result = response.text
    except Exception as e:
        logger.error(f'Error generating content: {e}')
        result = 'Error generating content'

    # Add the result to the Excel file
    if os.path.exists(EXCEL_FILE):
        wb = load_workbook(EXCEL_FILE)
    else:
        wb = Workbook()
        wb.active.append(["Username", "Result"])  # Add headers if it's a new file

    ws = wb.active
    ws.append([update.message.from_user.username, result])
    wb.save(EXCEL_FILE)

    await update.message.reply_text(f'Result added to Excel file: {result}')

# Function to handle errors
async def error(update: Update, context: ContextTypes.DEFAULT_TYPE):
    logger.warning(f'Update {update} caused error {context.error}')

# Main function
def main():
    application = Application.builder().token(TELEGRAM_TOKEN).build()

    application.add_handler(CommandHandler("start", start))
    application.add_handler(MessageHandler(filters.PHOTO, handle_photo))
    application.add_error_handler(error)

    application.run_polling()

if __name__ == '__main__':
    main()
